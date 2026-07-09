#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
領収書仕分け表 → 商工会MA1(会計王25) 仕訳データ受入用 .xls 変換スクリプト

使い方:
  python3 convert_to_ma1.py 入力.xlsx 出力_MA1取込用.xls --year 2026

入力(xlsx/csv)に必要な列（表記ゆれ可）:
  日付 / 店名(店名／支払先) / 金額(金額（円）) / 勘定科目 / 摘要(摘要・備考)
必要: pip install openpyxl xlwt xlrd
"""
import sys, argparse, datetime, os

# ---- 勘定科目コード＋既定税率（会計王25。要に応じて編集） ----
KAMOKU = {
    '消耗品費':      (575, '消耗品費',       '10%'),
    '仕入高':        (533, '仕入（8％）',     '8%軽'),   # 食材=軽減。酒/飲料は532へ手直し
    '仕入（8％）':   (533, '仕入（8％）',     '8%軽'),
    '仕入（10％）':  (532, '仕入（10％）',    '10%'),
    '旅費交通費':    (567, '旅費交通費',      '10%'),
    '接待交際費':    (571, '接待交際費（8％）','10%'),
    '会議費':        (600, '会議費',          '10%'),
    '修繕費':        (574, '修繕費',          '10%'),
    '支払手数料':    (581, '支払手数料',      '10%'),
    'リース料':      (583, 'リース料',        '10%'),
    '雑費':          (587, '雑費',            '10%'),
    # 同名コード未確認（暫定割当・要確認）
    '諸会費':        (587, '雑費',            '10%'),
    '福利厚生費':    (571, '接待交際費（8％）','8%軽'),
    '衛生費':        (587, '雑費',            '10%'),
    '水道光熱費':    (587, '雑費',            '10%'),
}
UNCONFIRMED = {'諸会費','福利厚生費','衛生費','水道光熱費'}
MIHARAI = (317, '未払金')   # カード払いの貸方（MA1実データ準拠）
CARD_WORDS = ('カード','クレジット','ｸﾚｼﾞｯﾄ','visa','VISA','ﾋﾞｻﾞ')
TEXT = {2,4,6,8,12,13,17,19,21,25,26,29,30,31}   # テキスト項目の列番号

def cap(s, nb):
    s = str(s).replace(',', '')
    b = s.encode('cp932', 'replace')
    return b[:nb].decode('cp932', 'ignore') if len(b) > nb else b.decode('cp932','ignore')

def to_reiwa(val, year_default):
    s = str(val).strip().replace('月','/').replace('日','').replace('年','/')
    s = s.replace('.', '/')
    parts = [p for p in s.replace('/',' ').split() if p != '']
    if len(parts) >= 3:
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2]); miss=False
    elif len(parts) == 2:
        y, m = int(parts[0]), int(parts[1]); d = 31 if m in (1,3,5,7,8,10,12) else 30; miss=True
    else:
        y, m, d = year_default, 1, 1; miss=True
    if y < 100: y += 2000
    return 'R.%02d/%02d/%02d' % (y-2018, m, d), miss

def split_tax(total, rate):
    div = 1.08 if rate == '8%軽' else 1.10
    h = int(round(total/div)); return h, total-h

def load_rows(path):
    rows=[]
    if path.lower().endswith('.csv'):
        import csv
        with open(path, encoding='utf-8-sig') as f:
            for r in csv.DictReader(f): rows.append(r)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # ヘッダ行を探す
        head=None; hi=0
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            vals=[str(c) if c is not None else '' for c in row]
            if any('日付' in v for v in vals) and any('金額' in v for v in vals):
                head=vals; hi=i; break
        if head is None: raise SystemExit('日付・金額の見出し行が見つかりません')
        for row in ws.iter_rows(min_row=hi+2, values_only=True):
            if all(c is None for c in row): continue
            rows.append({head[j]: row[j] for j in range(len(head)) if j < len(row)})
    return rows

def pick(d, *names):
    for k in d:
        kk=str(k)
        for n in names:
            if n in kk: return d[k]
    return None

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('infile'); ap.add_argument('outfile'); ap.add_argument('--year', type=int, default=datetime.date.today().year)
    a=ap.parse_args()
    import xlwt
    raw=load_rows(a.infile)
    now=datetime.datetime.now()
    head='//SFD FMTNAME="SIWAKE" FMTVER="3.04.00" APPNAME="会計王２５" APPVER="25.01.00" RTNDATE="%s" RTNTIME="%s"'%(now.strftime('%Y%m%d'),now.strftime('%H%M%S'))
    book=xlwt.Workbook(encoding='utf-8'); sh=book.add_sheet('Sheet1'); sh.write(0,0,head)
    notes=[]; den=0; total=0
    for r in raw:
        date=pick(r,'日付'); store=pick(r,'店名','支払先','支払') or ''
        amtv=pick(r,'金額'); kamoku=str(pick(r,'勘定科目') or '').strip()
        memo=pick(r,'摘要','備考') or ''
        pay=str(pick(r,'支払方法','貸方','決済') or '')
        if amtv in (None,'','合計'): continue
        if '合計' in str(store) or str(kamoku).strip()=='': continue  # 合計行・科目なし行は除外
        try: amt=int(float(str(amtv).replace(',','').replace('円','')))
        except: continue
        if kamoku not in KAMOKU:
            notes.append(f'{store} {amt}円: 勘定科目「{kamoku}」が未登録。雑費(587)で暫定計上。'); code,name,rate=(587,'雑費','10%')
        else:
            code,name,rate=KAMOKU[kamoku]
            if kamoku in UNCONFIRMED: notes.append(f'{store} {amt}円: 「{kamoku}」は暫定コード{code}。要確認。')
        rstr,miss=to_reiwa(date, a.year)
        if miss: notes.append(f'{store} {amt}円: 日付不明→暫定 {rstr}。要修正。')
        h,tax=split_tax(amt,rate); den+=1; total+=amt
        c=[('' if i in TEXT else 0) for i in range(35)]
        c[0]=den; c[1]=1; c[2]=rstr; c[3]=code; c[4]=name
        c[9]=21; c[11]=1; c[13]=rate; c[14]=h; c[15]=tax
        is_card = any(w in (pay+' '+str(store)+' '+str(memo)) for w in CARD_WORDS)
        if is_card:
            c[16], c[17] = MIHARAI
            notes.append(f'{store} {amt}円: カード払い→貸方 未払金({MIHARAI[0]})。')
        else:
            c[16], c[17] = 111, '現金'
        c[24]=3; c[26]='0%'; c[27]=amt
        c[29]=cap(store,30); c[30]=cap(memo,30)     # 補助摘要は30バイト厳守
        for i in range(35):
            if i in TEXT: sh.row(den).set_cell_text(i, str(c[i]))
            else: sh.write(den, i, int(c[i]))
    book.save(a.outfile)
    # 確認メモ
    try:
        import openpyxl
        wb=openpyxl.Workbook(); ws=wb.active; ws.title='要確認'
        ws.append(['注意事項']); 
        for n in notes: ws.append([n])
        ws.column_dimensions['A'].width=90
        wb.save(os.path.splitext(a.outfile)[0]+'_確認メモ.xlsx')
    except Exception: pass
    print(f'完了: {den}件  合計{total:,}円 → {a.outfile}')
    if notes:
        print('--- 要確認 ---'); [print(' -',n) for n in notes]

if __name__=='__main__':
    main()
